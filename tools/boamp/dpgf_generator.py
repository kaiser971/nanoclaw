"""Génération du DPGF (Décomposition du Prix Global et Forfaitaire) au format Excel (.xlsx).

Produit un DPGF structuré pour les réponses aux appels d'offres, avec sections
forfait de base, prestations unitaires, prestations complémentaires,
transition et récapitulatif avec formules de calcul.

Offer types supportés : TMA, DEVELOPPEMENT, FORMATION, IA.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

from entreprise import ENTREPRISE as E


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_BLUE_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
_DARK_BLUE_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
_LIGHT_GREY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_FONT_TITLE = Font(name="Arial", size=16, bold=True, color="1F497D")
_FONT_SUBTITLE = Font(name="Arial", size=12, bold=True, color="1F497D")
_FONT_SECTION = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_FONT_HEADER = Font(name="Arial", size=10, bold=True)
_FONT_NORMAL = Font(name="Arial", size=10)
_FONT_SMALL = Font(name="Arial", size=9, italic=True, color="808080")
_FONT_TOTAL = Font(name="Arial", size=11, bold=True, color="1F497D")
_FONT_GRAND_TOTAL = Font(name="Arial", size=12, bold=True, color="1F497D")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_CURRENCY_FMT = '#,##0.00 €'


# ---------------------------------------------------------------------------
# DPGF configurations per offer type
# ---------------------------------------------------------------------------

DPGF_CONFIGS = {
    "TMA": {
        "forfait": [
            ("F01", "Maintenance corrective", "Forfait/mois", 12),
            ("F02", "Maintenance évolutive", "Jours/homme", ""),
            ("F03", "Maintenance préventive", "Forfait/mois", 12),
            ("F04", "Support utilisateur N2/N3", "Forfait/mois", 12),
        ],
        "profiles": [
            ("P01", "Directeur de projet"),
            ("P02", "Chef de projet TMA"),
            ("P03", "Architecte technique"),
            ("P04", "Développeur senior"),
            ("P05", "Développeur"),
            ("P06", "Développeur junior"),
            ("P07", "Testeur / QA"),
            ("P08", "Analyste fonctionnel"),
        ],
        "complementary": [
            ("C01", "Astreinte heures ouvrées", "Forfait/mois"),
            ("C02", "Astreinte HNO", "Forfait/mois"),
            ("C03", "Intervention urgente HNO", "Heure"),
            ("C04", "Formation / transfert", "Jour"),
            ("C05", "Audit de code / sécurité", "Forfait"),
        ],
        "transition": [
            ("T01", "Reprise documentaire", "Jour"),
            ("T02", "Transfert de connaissances", "Jour"),
            ("T03", "Tuilage prestataire sortant", "Jour"),
        ],
    },
    "DEVELOPPEMENT": {
        "forfait": [
            ("F01", "Phase de cadrage / discovery", "Forfait", 1),
            ("F02", "Conception UX/UI", "Jours/homme", ""),
            ("F03", "Développement frontend", "Jours/homme", ""),
            ("F04", "Développement backend", "Jours/homme", ""),
            ("F05", "Tests et recette", "Jours/homme", ""),
            ("F06", "Mise en production", "Forfait", 1),
        ],
        "profiles": [
            ("P01", "Directeur de projet"),
            ("P02", "Chef de projet / Scrum Master"),
            ("P03", "UX/UI Designer"),
            ("P04", "Développeur fullstack senior"),
            ("P05", "Développeur frontend"),
            ("P06", "Développeur backend"),
            ("P07", "DevOps"),
            ("P08", "Testeur / QA"),
        ],
        "complementary": [
            ("C01", "Hébergement / infogérance", "Forfait/mois"),
            ("C02", "Maintenance corrective post-livraison", "Forfait/mois"),
            ("C03", "Support utilisateur", "Forfait/mois"),
            ("C04", "Évolutions mineures", "Jours/homme"),
            ("C05", "Audit accessibilité RGAA", "Forfait"),
        ],
        "transition": [
            ("T01", "Cadrage et spécifications", "Jour"),
            ("T02", "Prototype / PoC", "Jour"),
            ("T03", "Setup environnements", "Jour"),
        ],
    },
    "FORMATION": {
        "forfait": [
            ("F01", "Ingénierie pédagogique", "Jours/homme", ""),
            ("F02", "Création de contenu e-learning", "Module", ""),
            ("F03", "Configuration plateforme LMS", "Forfait", 1),
            ("F04", "Animation sessions (présentiel/distanciel)", "Jour", ""),
            ("F05", "Évaluation et certification", "Forfait", ""),
        ],
        "profiles": [
            ("P01", "Directeur de projet"),
            ("P02", "Ingénieur pédagogique"),
            ("P03", "Formateur expert web"),
            ("P04", "Développeur e-learning"),
            ("P05", "Graphiste / motion designer"),
            ("P06", "Intégrateur LMS"),
        ],
        "complementary": [
            ("C01", "Session présentiel supplémentaire", "Jour"),
            ("C02", "Mise à jour de contenu", "Module"),
            ("C03", "Support plateforme LMS", "Forfait/mois"),
            ("C04", "Traduction contenu", "Module"),
            ("C05", "Rapport d'évaluation", "Forfait"),
        ],
        "transition": [
            ("T01", "Analyse des besoins", "Jour"),
            ("T02", "Benchmark et choix LMS", "Jour"),
            ("T03", "Formation des formateurs", "Jour"),
        ],
    },
    "IA": {
        "forfait": [
            ("F01", "Audit et qualification des données", "Forfait", 1),
            ("F02", "Cadrage et PoC", "Jours/homme", ""),
            ("F03", "Développement modèle IA", "Jours/homme", ""),
            ("F04", "Intégration et API", "Jours/homme", ""),
            ("F05", "Tests et validation", "Jours/homme", ""),
            ("F06", "Mise en production et monitoring", "Forfait", 1),
        ],
        "profiles": [
            ("P01", "Directeur de projet IA"),
            ("P02", "Data Scientist senior"),
            ("P03", "ML Engineer"),
            ("P04", "Data Engineer"),
            ("P05", "Développeur intégration"),
            ("P06", "Responsable éthique IA"),
        ],
        "complementary": [
            ("C01", "Ré-entraînement modèle", "Forfait"),
            ("C02", "Extension périmètre données", "Jour"),
            ("C03", "Audit éthique et biais", "Forfait"),
            ("C04", "Monitoring et alerting", "Forfait/mois"),
            ("C05", "Formation équipe interne", "Jour"),
        ],
        "transition": [
            ("T01", "Audit données existantes", "Jour"),
            ("T02", "Proof of Concept", "Jour"),
            ("T03", "Setup MLOps", "Jour"),
        ],
    },
}

# Section titles per offer type
_SECTION_TITLES = {
    "TMA": {
        "s1": "Forfait de base (maintenance)",
        "s2": "Prestations unitaires (TJM)",
        "s3": "Prestations complémentaires",
        "s4": "Phase de transition",
    },
    "DEVELOPPEMENT": {
        "s1": "Phases de développement",
        "s2": "Prestations unitaires (TJM)",
        "s3": "Services complémentaires",
        "s4": "Phase de cadrage",
    },
    "FORMATION": {
        "s1": "Prestations pédagogiques",
        "s2": "Prestations unitaires (TJM)",
        "s3": "Services complémentaires",
        "s4": "Phase d'analyse",
    },
    "IA": {
        "s1": "Phases du projet IA",
        "s2": "Prestations unitaires (TJM)",
        "s3": "Services complémentaires",
        "s4": "Phase d'audit",
    },
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_cell(ws, row: int, col: int, value=None, font=None,
                fill=None, alignment=None, border=None,
                number_format=None) -> None:
    """Apply style to a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _write_section_header(ws, row: int, text: str, col_count: int) -> int:
    """Write a dark-blue section header spanning all columns. Returns next row."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=col_count,
    )
    _style_cell(ws, row, 1, value=text,
                font=_FONT_SECTION, fill=_DARK_BLUE_FILL,
                alignment=_ALIGN_LEFT, border=_THIN_BORDER)
    for c in range(2, col_count + 1):
        _style_cell(ws, row, c, font=_FONT_SECTION,
                     fill=_DARK_BLUE_FILL, border=_THIN_BORDER)
    return row + 1


def _write_table_header(ws, row: int, headers: list[str]) -> int:
    """Write a light-blue table header row. Returns next row."""
    for col, header in enumerate(headers, start=1):
        _style_cell(ws, row, col, value=header,
                     font=_FONT_HEADER, fill=_BLUE_FILL,
                     alignment=_ALIGN_CENTER, border=_THIN_BORDER)
    return row + 1


def _write_data_row(ws, row: int, values: list, col_count: int,
                    currency_cols: set[int] | None = None,
                    formula_cols: dict[int, str] | None = None,
                    editable_cols: set[int] | None = None) -> int:
    """Write a data row. Returns next row."""
    currency_cols = currency_cols or set()
    formula_cols = formula_cols or {}
    editable_cols = editable_cols or set()

    for col in range(1, col_count + 1):
        val = values[col - 1] if col - 1 < len(values) else ""
        fmt = _CURRENCY_FMT if col in currency_cols else None
        fill = _YELLOW_FILL if col in editable_cols else None
        align = _ALIGN_RIGHT if col in currency_cols else _ALIGN_LEFT

        if col in formula_cols:
            _style_cell(ws, row, col, value=formula_cols[col],
                         font=_FONT_NORMAL, fill=_LIGHT_GREY_FILL,
                         alignment=_ALIGN_RIGHT, border=_THIN_BORDER,
                         number_format=_CURRENCY_FMT)
        else:
            _style_cell(ws, row, col, value=val,
                         font=_FONT_NORMAL, fill=fill,
                         alignment=align, border=_THIN_BORDER,
                         number_format=fmt)
    return row + 1


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_dpgf(notice: dict[str, Any], dest_path: Path,
                  offer_type: str = "TMA") -> Path:
    """Generate a DPGF Excel workbook for a BOAMP notice.

    Args:
        notice: BOAMP notice data dict.
        dest_path: Path to save the .xlsx file.
        offer_type: Type of offer — "TMA", "DEVELOPPEMENT", "FORMATION", or "IA".

    Returns:
        Path to the generated file.
    """
    offer_type = offer_type.upper()
    if offer_type not in DPGF_CONFIGS:
        raise ValueError(
            f"Unknown offer_type {offer_type!r}. "
            f"Expected one of: {', '.join(DPGF_CONFIGS)}"
        )

    config = DPGF_CONFIGS[offer_type]
    titles = _SECTION_TITLES[offer_type]

    objet = notice.get("objet", "[OBJET DU MARCHÉ]")
    idweb = notice.get("idweb", "[REF]")

    wb = Workbook()
    ws = wb.active
    ws.title = "DPGF"

    # Column widths
    col_widths = [10, 42, 18, 16, 18, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    NUM_COLS = 6
    row = 1

    # ── Document header ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    _style_cell(ws, row, 1, value=E["raison_sociale"],
                font=_FONT_TITLE, alignment=_ALIGN_CENTER)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    _style_cell(ws, row, 1, value=f"DPGF — {objet}",
                font=_FONT_SUBTITLE, alignment=_ALIGN_CENTER)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    _style_cell(ws, row, 1, value=f"Référence BOAMP : {idweb}",
                font=_FONT_SMALL, alignment=_ALIGN_CENTER)
    row += 2

    # Track section total rows for the recap
    section_total_rows: dict[str, int] = {}
    f_col = get_column_letter(6)

    # ==================================================================
    # SECTION 1 — Forfait / Phases
    # ==================================================================
    row = _write_section_header(
        ws, row,
        f"Section 1 — {titles['s1']}",
        NUM_COLS,
    )
    headers_s1 = ["Réf", "Désignation", "Unité", "Qté estimée",
                  "PU HT (€)", "Total HT (€)"]
    row = _write_table_header(ws, row, headers_s1)

    s1_start = row
    for ref, desc, unite, qty in config["forfait"]:
        d_col = get_column_letter(4)
        e_col = get_column_letter(5)
        formula = f"=IF({e_col}{row}<>\"\",{d_col}{row}*{e_col}{row},\"\")"
        row = _write_data_row(
            ws, row, [ref, desc, unite, qty, "", ""],
            col_count=NUM_COLS,
            currency_cols={5, 6},
            editable_cols={5},
            formula_cols={6: formula},
        )

    s1_end = row - 1
    # Section subtotal
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _style_cell(ws, row, 1,
                value=f"Sous-total Section 1 — {titles['s1']}",
                font=_FONT_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER)
    for c in range(2, 6):
        _style_cell(ws, row, c, fill=_BLUE_FILL, border=_THIN_BORDER)
    _style_cell(ws, row, 6,
                value=f"=SUM({f_col}{s1_start}:{f_col}{s1_end})",
                font=_FONT_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER,
                number_format=_CURRENCY_FMT)
    section_total_rows["s1"] = row
    row += 2

    # ==================================================================
    # SECTION 2 — Prestations unitaires (TJM)
    # ==================================================================
    row = _write_section_header(
        ws, row,
        f"Section 2 — {titles['s2']}",
        NUM_COLS,
    )
    headers_s2 = ["Réf", "Profil", "", "", "PU HT (€/jour)", ""]
    row = _write_table_header(ws, row, headers_s2)
    # Merge header cells for cleaner look
    ws.merge_cells(start_row=row - 1, start_column=2,
                   end_row=row - 1, end_column=4)

    s2_start = row
    for ref, profil in config["profiles"]:
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=4)
        _style_cell(ws, row, 1, value=ref,
                     font=_FONT_NORMAL, alignment=_ALIGN_LEFT,
                     border=_THIN_BORDER)
        _style_cell(ws, row, 2, value=profil,
                     font=_FONT_NORMAL, alignment=_ALIGN_LEFT,
                     border=_THIN_BORDER)
        for c in range(3, 5):
            _style_cell(ws, row, c, border=_THIN_BORDER)
        _style_cell(ws, row, 5, font=_FONT_NORMAL,
                     fill=_YELLOW_FILL, alignment=_ALIGN_RIGHT,
                     border=_THIN_BORDER, number_format=_CURRENCY_FMT)
        _style_cell(ws, row, 6, border=_THIN_BORDER)
        row += 1

    s2_end = row - 1
    # No subtotal sum for TJM section (unit prices only)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _style_cell(ws, row, 1,
                value="Section 2 — Tarifs unitaires (à appliquer sur bons de commande)",
                font=_FONT_SMALL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER)
    for c in range(2, 7):
        _style_cell(ws, row, c, fill=_BLUE_FILL, border=_THIN_BORDER)
    section_total_rows["s2"] = row
    row += 2

    # ==================================================================
    # SECTION 3 — Prestations complémentaires
    # ==================================================================
    row = _write_section_header(
        ws, row,
        f"Section 3 — {titles['s3']}",
        NUM_COLS,
    )
    headers_s3 = ["Réf", "Désignation", "Unité", "Qté estimée",
                  "PU HT (€)", "Total HT (€)"]
    row = _write_table_header(ws, row, headers_s3)

    s3_start = row
    for ref, desc, unite in config["complementary"]:
        d_col = get_column_letter(4)
        e_col = get_column_letter(5)
        formula = f"=IF({e_col}{row}<>\"\",{d_col}{row}*{e_col}{row},\"\")"
        row = _write_data_row(
            ws, row, [ref, desc, unite, "", "", ""],
            col_count=NUM_COLS,
            currency_cols={5, 6},
            editable_cols={4, 5},
            formula_cols={6: formula},
        )

    s3_end = row - 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _style_cell(ws, row, 1,
                value=f"Sous-total Section 3 — {titles['s3']}",
                font=_FONT_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER)
    for c in range(2, 6):
        _style_cell(ws, row, c, fill=_BLUE_FILL, border=_THIN_BORDER)
    _style_cell(ws, row, 6,
                value=f"=SUM({f_col}{s3_start}:{f_col}{s3_end})",
                font=_FONT_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER,
                number_format=_CURRENCY_FMT)
    section_total_rows["s3"] = row
    row += 2

    # ==================================================================
    # SECTION 4 — Transition
    # ==================================================================
    row = _write_section_header(
        ws, row,
        f"Section 4 — {titles['s4']}",
        NUM_COLS,
    )
    headers_s4 = ["Réf", "Désignation", "Unité", "Qté estimée",
                  "PU HT (€)", "Total HT (€)"]
    row = _write_table_header(ws, row, headers_s4)

    s4_start = row
    for ref, desc, unite in config["transition"]:
        d_col = get_column_letter(4)
        e_col = get_column_letter(5)
        formula = f"=IF({e_col}{row}<>\"\",{d_col}{row}*{e_col}{row},\"\")"
        row = _write_data_row(
            ws, row, [ref, desc, unite, "", "", ""],
            col_count=NUM_COLS,
            currency_cols={5, 6},
            editable_cols={4, 5},
            formula_cols={6: formula},
        )

    s4_end = row - 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _style_cell(ws, row, 1,
                value=f"Sous-total Section 4 — {titles['s4']}",
                font=_FONT_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER)
    for c in range(2, 6):
        _style_cell(ws, row, c, fill=_BLUE_FILL, border=_THIN_BORDER)
    _style_cell(ws, row, 6,
                value=f"=SUM({f_col}{s4_start}:{f_col}{s4_end})",
                font=_FONT_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER,
                number_format=_CURRENCY_FMT)
    section_total_rows["s4"] = row
    row += 2

    # ==================================================================
    # SECTION 5 — Récapitulatif
    # ==================================================================
    row = _write_section_header(
        ws, row,
        "Section 5 — Récapitulatif",
        NUM_COLS,
    )

    recap_items = [
        (f"Sous-total {titles['s1']} (Section 1)", section_total_rows["s1"]),
        (f"Sous-total {titles['s3']} (Section 3)", section_total_rows["s3"]),
        (f"Sous-total {titles['s4']} (Section 4)", section_total_rows["s4"]),
    ]

    recap_start = row
    for label, total_row in recap_items:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=5)
        _style_cell(ws, row, 1, value=label,
                     font=_FONT_NORMAL, alignment=_ALIGN_RIGHT,
                     border=_THIN_BORDER)
        for c in range(2, 6):
            _style_cell(ws, row, c, border=_THIN_BORDER)
        _style_cell(ws, row, 6,
                     value=f"={f_col}{total_row}",
                     font=_FONT_NORMAL, alignment=_ALIGN_RIGHT,
                     border=_THIN_BORDER, number_format=_CURRENCY_FMT)
        row += 1

    recap_end = row - 1

    # Total HT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _style_cell(ws, row, 1, value="TOTAL HT",
                font=_FONT_GRAND_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER)
    for c in range(2, 6):
        _style_cell(ws, row, c, fill=_BLUE_FILL, border=_THIN_BORDER)
    total_ht_row = row
    _style_cell(ws, row, 6,
                value=f"=SUM({f_col}{recap_start}:{f_col}{recap_end})",
                font=_FONT_GRAND_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER,
                number_format=_CURRENCY_FMT)
    row += 1

    # TVA 20%
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _style_cell(ws, row, 1, value="TVA (20 %)",
                font=_FONT_TOTAL, alignment=_ALIGN_RIGHT,
                border=_THIN_BORDER)
    for c in range(2, 6):
        _style_cell(ws, row, c, border=_THIN_BORDER)
    _style_cell(ws, row, 6,
                value=f"={f_col}{total_ht_row}*0.20",
                font=_FONT_TOTAL, alignment=_ALIGN_RIGHT,
                border=_THIN_BORDER, number_format=_CURRENCY_FMT)
    tva_row = row
    row += 1

    # Total TTC
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _style_cell(ws, row, 1, value="TOTAL TTC",
                font=_FONT_GRAND_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER)
    for c in range(2, 6):
        _style_cell(ws, row, c, fill=_BLUE_FILL, border=_THIN_BORDER)
    _style_cell(ws, row, 6,
                value=f"={f_col}{total_ht_row}+{f_col}{tva_row}",
                font=_FONT_GRAND_TOTAL, fill=_BLUE_FILL,
                alignment=_ALIGN_RIGHT, border=_THIN_BORDER,
                number_format=_CURRENCY_FMT)
    row += 2

    # ── Notes ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    _style_cell(ws, row, 1,
                value="Les cellules jaunes sont à compléter par le candidat.",
                font=_FONT_SMALL, alignment=_ALIGN_LEFT)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    _style_cell(ws, row, 1,
                value="Section 2 (TJM) : tarifs unitaires appliqués sur bons de commande, non inclus dans le total forfaitaire.",
                font=_FONT_SMALL, alignment=_ALIGN_LEFT)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    _style_cell(ws, row, 1,
                value=f"{E['raison_sociale']} — {E['adresse']}, {E['code_postal']} {E['ville']} — SIRET {E['siret']}",
                font=_FONT_SMALL, alignment=_ALIGN_CENTER)

    # Print setup
    ws.print_title_rows = "1:4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"

    # ── Save ──
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest_path))
    return dest_path
